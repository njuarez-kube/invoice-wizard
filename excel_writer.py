"""
excel_writer.py
Reads and writes the Facturas Kube Excel file.
Only touches green columns; all other columns are left untouched.
"""

import datetime
import logging
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from processor import InvoiceData

# ── Green column indices (1-based) ────────────────────────────────────────────
COL_REF         = 1   # A  Ref
COL_INVOICE_NUM = 2   # B  Invoice
COL_DATE        = 3   # C  Date
COL_VAT_INC     = 6   # F  € (VAT inc)
COL_VAT_RETURN  = 7   # G  VAT Return.
COL_EXCL_VAT    = 8   # H  € (VAT excl)
COL_RETENTION   = 9   # I  Retención IRPF
# J (10) — left blank
COL_VENDOR      = 11  # K  To/From
# L (12) — left blank
COL_COMMENTS    = 13  # M  Comments (optional)

LAST_COL = COL_COMMENTS   # nothing written beyond here

GREEN_FILL = PatternFill("solid", fgColor="FF92D050")


def get_or_create_workbook(output_path: Path, template_path: Path) -> openpyxl.Workbook:
    if output_path.exists():
        return openpyxl.load_workbook(output_path)

    tmpl = openpyxl.load_workbook(template_path)
    tmpl_ws = tmpl.active

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tmpl_ws.title

    for cell in tmpl_ws[1]:
        if cell.column > COL_VENDOR:
            break
        nc = ws.cell(row=1, column=cell.column, value=cell.value)
        if cell.font:      nc.font      = copy(cell.font)
        if cell.fill:      nc.fill      = copy(cell.fill)
        if cell.border:    nc.border    = copy(cell.border)
        if cell.alignment: nc.alignment = copy(cell.alignment)
        nc.number_format = cell.number_format

    for col, name in [(COL_RETENTION, 'Retención'), (COL_COMMENTS, 'Comments')]:
        ws.cell(row=1, column=col, value=name).fill = GREEN_FILL

    for col_letter, dim in tmpl_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logging.info("Created new workbook: %s", output_path)
    return wb


def load_existing_invoice_numbers(output_path: Path) -> set[tuple[str, str]]:
    """Return a set of (invoice_number, vendor_name) pairs already in the workbook."""
    if not output_path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
        ws = wb.active
        seen: set[tuple[str, str]] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            num    = row[COL_INVOICE_NUM - 1]
            vendor = row[COL_VENDOR - 1]
            if num is not None:
                seen.add((str(num).strip(), str(vendor).strip() if vendor else ''))
        wb.close()
        return seen
    except Exception as exc:
        logging.warning("Could not read existing invoice numbers: %s", exc)
        return set()


def _find_next_empty_row(ws) -> int:
    row = 2
    while ws.cell(row=row, column=COL_INVOICE_NUM).value is not None:
        row += 1
    return row


def _write_row(ws, row: int, inv: InvoiceData) -> None:
    vat_inc = inv.vat_inc
    if vat_inc is None and inv.excl_vat is not None and inv.vat_amount is not None:
        vat_inc = round(inv.excl_vat + inv.vat_amount, 2)
    # When retention is present: € (VAT inc) = € (VAT excl) + VAT Return. - Retención
    if inv.retention is not None and inv.excl_vat is not None and inv.vat_amount is not None:
        vat_inc = round(inv.excl_vat + inv.vat_amount - inv.retention, 2)

    ref = Path(inv.source_file).stem.split(' ')[0] if inv.source_file else None

    cells_values = [
        (COL_REF,         ref),
        (COL_INVOICE_NUM, inv.invoice_number),
        (COL_DATE,        inv.date),
        (COL_VAT_INC,     vat_inc),
        (COL_VAT_RETURN,  inv.vat_amount),
        (COL_EXCL_VAT,    inv.excl_vat),
        (COL_RETENTION,   inv.retention),
        (COL_VENDOR,      inv.vendor_name),
        (COL_COMMENTS,    inv.comments),
    ]

    for col, value in cells_values:
        cell = ws.cell(row=row, column=col, value=value)
        if col == COL_DATE and isinstance(value, datetime.date):
            cell.number_format = 'DD/MM/YYYY'


def _sort_rows(ws) -> None:
    """Sort all data rows by column A (Ref) ascending; None values go last."""
    last_row = max(
        (r for r in range(2, ws.max_row + 1)
         if any(ws.cell(row=r, column=c).value is not None for c in range(1, LAST_COL + 1))),
        default=1,
    )
    if last_row < 2:
        return

    all_rows = [
        {c: (ws.cell(row=r, column=c).value, ws.cell(row=r, column=c).number_format)
         for c in range(1, LAST_COL + 1)}
        for r in range(2, last_row + 1)
    ]

    all_rows.sort(key=lambda row: (row[COL_REF][0] is None, str(row[COL_REF][0] or '').lower()))

    for i, row_data in enumerate(all_rows):
        for col, (value, fmt) in row_data.items():
            cell = ws.cell(row=i + 2, column=col, value=value)
            if fmt:
                cell.number_format = fmt


def append_rows(
    output_path: Path,
    template_path: Path,
    invoices: list[InvoiceData],
    seen: set[tuple[str, str]],
) -> dict:
    """
    Append invoice rows to the workbook. Returns a summary dict.
    Mutates `seen` in place so callers can chain multiple calls.
    """
    wb = get_or_create_workbook(output_path, template_path)
    ws = wb.active

    written = 0
    skipped_duplicates = 0
    errors = 0

    for inv in invoices:
        num    = inv.invoice_number or ''
        vendor = inv.vendor_name or ''
        key    = (num, vendor)
        if num and key in seen:
            logging.warning("Duplicate invoice skipped: %s (%s)", num, vendor)
            skipped_duplicates += 1
            continue

        if inv.invoice_number is None and inv.date is None and inv.excl_vat is None:
            logging.warning("Empty invoice data skipped (%s %s)", inv.source_file, inv.page_range)
            errors += 1
            continue

        row = _find_next_empty_row(ws)
        _write_row(ws, row, inv)
        if num:
            seen.add(key)
        written += 1

    _sort_rows(ws)
    wb.save(output_path)
    return {'written': written, 'duplicates': skipped_duplicates, 'errors': errors}
