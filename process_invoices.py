#!/usr/bin/env python3
"""
process_invoices.py
Batch-processes Amazon invoice PDFs and appends rows to output/gastos.xlsx.

Usage:
    python process_invoices.py [--template PATH] [--output PATH]

Folder layout (relative to this script):
    input/   <- drop PDFs here before running
    done/    <- processed PDFs are moved here automatically
    output/  <- gastos.xlsx is created/updated here
"""

import re
import shutil
import logging
import argparse
import configparser
import datetime
from pathlib import Path
from copy import copy

import pdfplumber
import openpyxl

# ── Column indices (1-based) ──────────────────────────────────────────────────
COL_INVOICE_NUM = 2   # B
COL_DATE        = 3   # C
COL_VAT_RETURN  = 7   # G  (VAT Returnable)
COL_EXCL_VAT    = 8   # H  (VAT excl)
COL_VENDOR      = 11  # K  (To/From)
COL_COMMENTS    = 13  # M  (Comments)

# ── Spanish month names ───────────────────────────────────────────────────────
SPANISH_MONTHS = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}

# ── Regex patterns ────────────────────────────────────────────────────────────
# "Número de la factura ES6009311LVPBI"
RE_INVOICE_NUM = re.compile(
    r'N[uú]mero\s+de\s+la\s+factura\s+([A-Z0-9]+)',
    re.IGNORECASE,
)
# "Fecha de la factura/Fecha de la entrega 30 abril 2026"
RE_DATE = re.compile(
    r'Fecha\s+de\s+la\s+factura[^\n]*?(\d{1,2})\s+([a-záéíóú]+)\s+(\d{4})',
    re.IGNORECASE,
)
# 3-column address header line followed by the data line
# "Dirección comercial Dirección de envío Vendido por\n[billing] [recipient] [vendor]"
RE_ADDR_LINE = re.compile(
    r'Direcci[oó]n\s+comercial\s+Direcci[oó]n\s+de\s+env[ií]o\s+Vendido\s+por\s*\n(.+)',
    re.IGNORECASE,
)

# VAT breakdown rows: "21% 5,78 € 1,21 €" or "0% 14,86 € 0,00 €"
RE_VAT_ROW = re.compile(
    r'^(\d+)%\s+([\d,]+)\s+\S+\s+([\d,]+)',
    re.MULTILINE,
)
# Fallback total: "Total pendiente 14,86 €"
RE_TOTAL_PENDING = re.compile(r'Total\s+pendiente\s+([\d,]+)', re.IGNORECASE)


# ── Utility ───────────────────────────────────────────────────────────────────
def parse_spanish_number(s: str) -> float | None:
    """Convert '14,86' → 14.86 and '1.234,56' → 1234.56."""
    if not s:
        return None
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except ValueError:
        return None


# ── Field parsers ─────────────────────────────────────────────────────────────
def parse_invoice_number(text: str) -> str | None:
    m = RE_INVOICE_NUM.search(text)
    if m:
        return m.group(1).strip()
    logging.warning("Could not extract invoice number")
    return None


def parse_date(text: str) -> datetime.date | None:
    m = RE_DATE.search(text)
    if m:
        day = int(m.group(1))
        month = SPANISH_MONTHS.get(m.group(2).lower())
        year = int(m.group(3))
        if month:
            try:
                return datetime.date(year, month, day)
            except ValueError:
                pass
    logging.warning("Could not extract date")
    return None


def parse_vendor(_text: str) -> str:
    return "Amazon"


def parse_recipient(text: str) -> str | None:
    """Extract the shipping recipient name from the 3-column address line."""
    m = RE_ADDR_LINE.search(text)
    if not m:
        logging.warning("Could not extract recipient name")
        return None
    line = m.group(1)
    # The line is: "[Billing S.L.] [Recipient Name] [Vendor name]"
    # Strip the vendor name from the right so it doesn't bleed into the match
    vendor_m = re.search(r'^Vendido\s+por\s+(.+)$', text, re.IGNORECASE | re.MULTILINE)
    if vendor_m:
        vendor_name = vendor_m.group(1).strip()
        idx = line.find(vendor_name)
        if idx > 0:
            line = line[:idx].strip()
    # Recipient name follows the billing company's legal suffix
    name_m = re.search(
        r'(?:S\.L\.|S\.A\.|SL\b|SA\b|Ltd\.?|GmbH|Inc\.?|Corp\.?|LLC)\s+(.+)',
        line,
        re.UNICODE,
    )
    if name_m:
        return name_m.group(1).strip()
    logging.warning("Could not extract recipient name from: %s", line)
    return None


def parse_amounts(text: str) -> tuple[float | None, float | None]:
    """
    Returns (excl_vat, vat_amount).
    Reads the VAT breakdown table rows (e.g. '21% 5,78 € 1,21 €').
    Falls back to Total pendiente with 0 VAT if no table rows found.
    """
    rows = RE_VAT_ROW.findall(text)
    if rows:
        excl_vat   = sum(parse_spanish_number(r[1]) or 0.0 for r in rows)
        vat_amount = sum(parse_spanish_number(r[2]) or 0.0 for r in rows)
        return round(excl_vat, 2), round(vat_amount, 2)

    m = RE_TOTAL_PENDING.search(text)
    if m:
        total = parse_spanish_number(m.group(1))
        if total is not None:
            return total, 0.0

    logging.warning("Could not extract amounts")
    return None, None


# ── Excel helpers ─────────────────────────────────────────────────────────────
def get_or_create_workbook(
    output_path: Path, template_path: Path
) -> openpyxl.Workbook:
    """Load existing workbook or bootstrap one from the template header row."""
    if output_path.exists():
        return openpyxl.load_workbook(output_path)

    tmpl = openpyxl.load_workbook(template_path)
    tmpl_ws = tmpl.active

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tmpl_ws.title

    for cell in tmpl_ws[1]:
        nc = ws.cell(row=1, column=cell.column, value=cell.value)
        if cell.font:       nc.font       = copy(cell.font)
        if cell.fill:       nc.fill       = copy(cell.fill)
        if cell.border:     nc.border     = copy(cell.border)
        if cell.alignment:  nc.alignment  = copy(cell.alignment)
        nc.number_format = cell.number_format

    for col_letter, dim in tmpl_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logging.info("Created new workbook: %s", output_path)
    return wb


def find_next_empty_row(ws) -> int:
    row = 2
    while ws.cell(row=row, column=COL_INVOICE_NUM).value is not None:
        row += 1
    return row


def append_invoice_row(
    ws,
    row: int,
    invoice_number: str | None,
    date: datetime.date | None,
    vat_returnable: float | None,
    excl_vat: float | None,
    vendor: str | None,
    product: str | None,
) -> None:
    ws.cell(row=row, column=COL_INVOICE_NUM).value = invoice_number
    date_cell = ws.cell(row=row, column=COL_DATE)
    date_cell.value = date
    if date is not None:
        date_cell.number_format = 'DD/MM/YYYY'
    ws.cell(row=row, column=COL_VAT_RETURN).value = vat_returnable
    ws.cell(row=row, column=COL_EXCL_VAT).value   = excl_vat
    ws.cell(row=row, column=COL_VENDOR).value      = vendor
    ws.cell(row=row, column=COL_COMMENTS).value    = product


# ── PDF pipeline ──────────────────────────────────────────────────────────────
def process_pdf(pdf_path: Path, ws) -> int:
    """
    Process all invoice pages in one PDF, append rows to ws.
    Each page that contains an invoice number is treated as one invoice.
    Returns the number of invoices successfully appended.
    """
    count = 0
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ''

                if not RE_INVOICE_NUM.search(text):
                    continue  # continuation page or non-invoice page

                invoice_number = parse_invoice_number(text)
                date           = parse_date(text)
                vendor         = parse_vendor(text)
                product        = parse_recipient(text)
                excl_vat, vat_amount = parse_amounts(text)

                row = find_next_empty_row(ws)
                append_invoice_row(
                    ws, row,
                    invoice_number, date, vat_amount, excl_vat, vendor, product
                )

                logging.info(
                    "  [p%d] %s | %s | %s | excl=%.2f | vat=%.2f",
                    page_num,
                    invoice_number or '?',
                    date or '?',
                    (vendor or '?')[:40],
                    excl_vat or 0.0,
                    vat_amount or 0.0,
                )
                count += 1

    except Exception as exc:
        logging.error("Error processing %s: %s", pdf_path.name, exc)
        return 0

    return count


# ── Config helpers ────────────────────────────────────────────────────────────
def load_config(script_dir: Path) -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg_path = script_dir / 'config.ini'
    if cfg_path.exists():
        cfg.read(cfg_path, encoding='utf-8')
    return cfg


def resolve_path(value: str, script_dir: Path) -> Path:
    """Return an absolute Path; relative paths are anchored to script_dir."""
    p = Path(value)
    return p if p.is_absolute() else script_dir / p


# ── Entry point ───────────────────────────────────────────────────────────────
def main() -> None:
    script_dir = Path(__file__).parent
    cfg = load_config(script_dir)
    paths = cfg['paths'] if 'paths' in cfg else {}

    parser = argparse.ArgumentParser(
        description="Batch-process Amazon invoice PDFs → Excel"
    )
    parser.add_argument(
        '--template',
        default=paths.get('template', str(script_dir / 'ejemplo.xlsx')),
        help='Path to the Excel template',
    )
    parser.add_argument(
        '--output',
        default=paths.get('output', str(script_dir / 'output' / 'gastos.xlsx')),
        help='Output Excel file path',
    )
    parser.add_argument(
        '--input-dir',
        default=paths.get('input_dir', str(script_dir / 'input')),
        help='Folder to read PDFs from',
    )
    parser.add_argument(
        '--done-dir',
        default=paths.get('done_dir', str(script_dir / 'done')),
        help='Folder to move processed PDFs to',
    )
    args = parser.parse_args()

    input_dir     = resolve_path(args.input_dir, script_dir)
    done_dir      = resolve_path(args.done_dir, script_dir)
    output_path   = resolve_path(args.output, script_dir)
    template_path = resolve_path(args.template, script_dir)

    input_dir.mkdir(exist_ok=True)
    done_dir.mkdir(exist_ok=True)
    output_path.parent.mkdir(exist_ok=True)

    pdfs = sorted(input_dir.glob('*.pdf'))
    if not pdfs:
        logging.info("No PDFs found in %s — nothing to do.", input_dir)
        return

    logging.info("Found %d PDF(s) in %s", len(pdfs), input_dir)

    wb = get_or_create_workbook(output_path, template_path)
    ws = wb.active

    total = 0
    for pdf_path in pdfs:
        logging.info("Processing: %s", pdf_path.name)
        n = process_pdf(pdf_path, ws)
        if n > 0:
            wb.save(output_path)
            dest = done_dir / pdf_path.name
            shutil.move(str(pdf_path), dest)
            logging.info("  -> %d invoice(s) saved, file moved to done/", n)
            total += n
        else:
            logging.warning("  -> No invoices extracted; file left in input/")

    logging.info("Finished. Total invoices added this run: %d", total)
    logging.info("Output: %s", output_path)


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    main()
